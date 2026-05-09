import pandas as pd
from openpyxl import load_workbook

# Load and normalize FRED data
fred = pd.read_csv("PCU4831114831115.csv", parse_dates=["observation_date"])
fred["year"] = fred["observation_date"].dt.year
fred["month"] = fred["observation_date"].dt.month
fred["value"] = fred["PCU4831114831115"]

base = fred[fred["year"] == 2025]["value"].mean()
fred["index"] = (fred["value"] / base * 100).round(1)

pivot = fred.pivot(index="year", columns="month", values="index")

# Write to Excel
wb = load_workbook("data.xlsx")
ws = wb.active

next_row = ws.max_row + 2

ws.cell(row=next_row, column=1, value="Global Freight PPI (FRED)")
next_row += 1
ws.cell(row=next_row, column=1, value="Year")
for m, name in enumerate(["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"], 1):
    ws.cell(row=next_row, column=m + 1, value=name)
next_row += 1

for year, row in pivot.iterrows():
    ws.cell(row=next_row, column=1, value=year)
    for m in range(1, 13):
        val = row.get(m)
        if pd.notna(val):
            ws.cell(row=next_row, column=m + 1, value=val)
    next_row += 1

wb.save("data.xlsx")
print("Done — data.xlsx updated")